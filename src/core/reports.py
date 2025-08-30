"""
Модуль генерации отчетов о процессе загрузки
"""
import os
import csv
from datetime import datetime
from typing import Dict, List
from openpyxl import Workbook


def generate_upload_report(grouped, upload_results: Dict[str, List[str]], warnings: List[str] = None, export_path: str = None):
    """
    Генерирует подробный отчет о процессе загрузки
    
    Args:
        grouped: Результат парсинга фото (объект с by_sku)
        upload_results: Словарь sku -> список ссылок
        warnings: Список предупреждений
        export_path: Путь для сохранения отчета
    
    Returns:
        str: Путь к созданному отчету
    """
    if not export_path:
        timestamp = datetime.now().strftime('%Y%m%d-%H%M%S')
        export_path = f'wb_upload_report_{timestamp}.xlsx'
    
    wb = Workbook()
    
    # Лист 1: Сводка по SKU
    ws_summary = wb.active
    ws_summary.title = "Сводка по SKU"
    
    headers = [
        'SKU', 'Найдено файлов', 'Загружено ссылок', 'Статус', 
        'Локальные файлы', 'Публичные ссылки'
    ]
    ws_summary.append(headers)
    
    for sku, files in (grouped.by_sku.items() if grouped else {}):
        local_files = [os.path.basename(f.path) for f in files]
        uploaded_links = upload_results.get(sku, [])
        
        status = "✅ Успешно" if uploaded_links else "❌ Ошибка"
        if len(uploaded_links) < len(files):
            status = "⚠️ Частично"
            
        row = [
            sku,
            len(files),
            len(uploaded_links),
            status,
            '; '.join(local_files),
            '; '.join(uploaded_links)
        ]
        ws_summary.append(row)
    
    # Лист 2: Предупреждения
    if warnings:
        ws_warnings = wb.create_sheet("Предупреждения")
        ws_warnings.append(['Тип', 'Сообщение'])
        
        for warning in warnings:
            ws_warnings.append(['Предупреждение', warning])
    
    # Лист 3: Статистика
    ws_stats = wb.create_sheet("Статистика")
    total_sku = len(grouped.by_sku) if grouped else 0
    total_files = sum(len(files) for files in (grouped.by_sku.values() if grouped else []))
    total_links = sum(len(links) for links in upload_results.values())
    success_sku = len([sku for sku, links in upload_results.items() if links])
    
    stats = [
        ['Параметр', 'Значение'],
        ['Время генерации', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ['Всего SKU', total_sku],
        ['Всего файлов', total_files],
        ['Успешно загружено ссылок', total_links],
        ['Успешных SKU', success_sku],
        ['Процент успеха', f'{(success_sku/max(1, total_sku)*100):.1f}%'],
        ['Среднее фото на SKU', f'{(total_files/max(1, total_sku)):.1f}'],
    ]
    
    for row in stats:
        ws_stats.append(row)
    
    # Сохранение
    wb.save(export_path)
    return export_path


def export_csv_report(grouped, upload_results: Dict[str, List[str]], export_path: str = None):
    """
    Экспортирует упрощенный отчет в CSV формате
    """
    if not export_path:
        timestamp = datetime.now().strftime('%Y%m%d-%H%M%S')
        export_path = f'wb_upload_report_{timestamp}.csv'
    
    with open(export_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['SKU', 'Файлов', 'Ссылок', 'Статус', 'Ссылки'])
        
        for sku, files in (grouped.by_sku.items() if grouped else {}):
            uploaded_links = upload_results.get(sku, [])
            status = "OK" if uploaded_links else "ERROR"
            
            writer.writerow([
                sku,
                len(files),
                len(uploaded_links),
                status,
                '; '.join(uploaded_links)
            ])
    
    return export_path
